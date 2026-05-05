import os
import sys
import json
import psycopg2
from psycopg2 import pool as pg_pool
from psycopg2.extras import RealDictCursor
from flask import Flask, request, jsonify, send_file, make_response
from flask_cors import CORS
import pandas as pd
from datetime import datetime
import io
import traceback
from pathlib import Path

app = Flask(__name__)
CORS(app)
app.config['JSON_SORT_KEYS'] = False

PORT = int(os.environ.get('PORT', 3000))

# ============================================================
# PATH CONFIGURATION
# ============================================================
IS_PACKAGED = getattr(sys, 'frozen', False)
EXECUTABLE_PATH = os.path.dirname(sys.executable) if IS_PACKAGED else os.path.dirname(os.path.abspath(__file__))
SNAPSHOT_PATH = os.path.dirname(os.path.abspath(__file__))

print(f"📦 Packaged mode: {'YES' if IS_PACKAGED else 'NO'}")
print(f"📁 Executable path: {EXECUTABLE_PATH}")

# ============================================================
# FIND FRONTEND FILES
# ============================================================
possible_frontend_paths = [
    os.path.join(SNAPSHOT_PATH, 'frontend'),
    os.path.join(EXECUTABLE_PATH, 'frontend'),
    os.path.join(os.path.dirname(os.path.abspath(__file__)), '../frontend'),
    os.path.join(os.path.dirname(os.path.abspath(__file__)), 'frontend'),
    os.path.join(os.getcwd(), 'frontend')
]

frontend_path = None
for test_path in possible_frontend_paths:
    indexPath = os.path.join(test_path, 'index.html')
    if os.path.exists(indexPath):
        frontend_path = test_path
        print(f"✅ Frontend found at: {test_path}")
        app.static_folder = frontend_path
        app.static_url_path = ''
        break

# ============================================================
# DATABASE CONNECTION
# ============================================================
connection_pool = None
db_config = None

# ============================================================
# LOAD CONFIGURATION
# ============================================================
def load_config():
    global db_config
    try:
        possible_config_paths = [
            os.path.join(EXECUTABLE_PATH, 'config.json'),
            os.path.join(SNAPSHOT_PATH, 'config.default.json'),
            os.path.join(os.path.dirname(os.path.abspath(__file__)), 'config.json'),
            os.path.join(os.getcwd(), 'config.json')
        ]

        for config_path in possible_config_paths:
            if os.path.exists(config_path):
                with open(config_path, 'r') as f:
                    db_config = json.load(f)
                print(f"✅ Config loaded from: {config_path}")
                print(f"📊 Database: {db_config.get('database', {}).get('database')}")
                return True

        print("⚠️ No config file found")
        return False
    except Exception as error:
        print(f"❌ Failed to load config: {str(error)}")
        return False

# ============================================================
# LOAD QUERIES
# ============================================================
def load_radet_query():
    """Load RADET report query"""
    try:
        possible_query_paths = [
            os.path.join(EXECUTABLE_PATH, 'queries', 'radet.sql'),
            os.path.join(SNAPSHOT_PATH, 'queries', 'radet.sql'),
            os.path.join(EXECUTABLE_PATH, 'radet_query.sql'),
            os.path.join(SNAPSHOT_PATH, 'radet_query.sql'),
            os.path.join(os.path.dirname(os.path.abspath(__file__)), 'radet_query.sql'),
            os.path.join(os.getcwd(), 'radet_query.sql')
        ]

        for query_path in possible_query_paths:
            if os.path.exists(query_path):
                print(f"📁 Using RADET query file: {query_path}")
                with open(query_path, 'r') as f:
                    return f.read()

        print("❌ RADET query file not found")
        return None
    except Exception as error:
        print(f"❌ Failed to load RADET query: {str(error)}")
        return None

def load_pharmacy_query():
    """Load Pharmacy report query"""
    try:
        possible_query_paths = [
            os.path.join(EXECUTABLE_PATH, 'queries', 'pharmacy.sql'),
            os.path.join(SNAPSHOT_PATH, 'queries', 'pharmacy.sql'),
            os.path.join(EXECUTABLE_PATH, 'pharmacy_query.sql'),
            os.path.join(SNAPSHOT_PATH, 'pharmacy_query.sql'),
            os.path.join(os.path.dirname(os.path.abspath(__file__)), 'pharmacy_query.sql'),
            os.path.join(os.getcwd(), 'pharmacy_query.sql')
        ]

        for query_path in possible_query_paths:
            if os.path.exists(query_path):
                print(f"📁 Using Pharmacy query file: {query_path}")
                with open(query_path, 'r') as f:
                    return f.read()

        print("❌ Pharmacy query file not found")
        return None
    except Exception as error:
        print(f"❌ Failed to load Pharmacy query: {str(error)}")
        return None

# ============================================================
# CONNECT TO DATABASE
# ============================================================
def connect_to_database():
    global connection_pool
    if not db_config or not db_config.get('database', {}).get('database'):
        print("⚠️ Database not configured")
        return False

    try:
        db = db_config['database']
        connection_pool = pg_pool.SimpleConnectionPool(
            1, 20,
            host=db.get('host'),
            port=db.get('port'),
            database=db.get('database'),
            user=db.get('username'),
            password=db.get('password'),
            sslmode='require' if db.get('ssl') else 'disable',
            connect_timeout=5
        )

        # Test connection
        conn = connection_pool.getconn()
        cursor = conn.cursor()
        cursor.execute('SELECT NOW()')
        cursor.close()
        connection_pool.putconn(conn)
        
        print(f"✅ Database connected successfully to {db.get('database')}")
        return True
    except Exception as error:
        print(f"❌ Database connection failed: {str(error)}")
        connection_pool = None
        return False

# ============================================================
# EXCEL GENERATION FUNCTIONS
# ============================================================

def generate_radet_excel(rows):
    """Generate RADET report Excel with 81 columns"""
    columns = [
        'S/No.', 'State', 'L.G.A', 'LGA Of Residence', 'Facility Name',
        'DatimId', 'Patient ID', 'NDR Patient Identifier', 'Hospital Number',
        'Unique Id', 'Household Unique No', 'OVC Unique ID', 'Sex',
        'Target group', 'Current Weight (kg)', 'Pregnancy Status',
        'Date of Birth (yyyy-mm-dd)', 'Age', 'Care Entry Point',
        'Date of Registration', 'Enrollment Date (yyyy-mm-dd)',
        'ART Start Date (yyyy-mm-dd)', 'Last Pickup Date (yyyy-mm-dd)',
        'Months of ARV Refill', 'Regimen Line at ART Start',
        'Regimen at ART Start', 'Date of Start of Current ART Regimen',
        'Current Regimen Line', 'Current ART Regimen',
        'Clinical Staging at Last Visit', 'Date of Last CD4 Count',
        'Last CD4 Count', 'Date of Viral Load Sample Collection (yyyy-mm-dd)',
        'Date of Current ViralLoad Result Sample (yyyy-mm-dd)',
        'Current Viral Load (c/ml)', 'Date of Current Viral Load (yyyy-mm-dd)',
        'Viral Load Indication', 'Viral Load Eligibility Status',
        'Date of Viral Load Eligibility Status', 'Current ART Status',
        'Date of Current ART Status', 'Client Verification Outcome',
        'Cause of Death', 'VA Cause of Death', 'Previous ART Status',
        'Confirmed Date of Previous ART Status', 'ART Enrollment Setting',
        'Date of TB Screening (yyyy-mm-dd)', 'TB Screening Type', 'CAD Score',
        'TB status', 'Date of TB Sample Collection (yyyy-mm-dd)',
        'TB Diagnostic Test Type', 'Date of TB Diagnostic Result Received (yyyy-mm-dd)',
        'TB Diagnostic Result', 'Date of Additional TB Diagnosis Result using XRAY',
        'Additional TB Diagnosis Result using XRAY',
        'Date of Start of TB Treatment (yyyy-mm-dd)',
        'TB Type (new, relapsed etc)', 'Date of Completion of TB Treatment (yyyy-mm-dd)',
        'TB Treatment Outcome', 'Date of TPT Start (yyyy-mm-dd)', 'TPT Type',
        'TPT Completion date (yyyy-mm-dd)', 'TPT Completion status',
        'Date of commencement of EAC (yyyy-mm-dd)',
        'Number of EAC Sessions Completed', 'Date of last EAC Session Completed',
        'Date of Extended EAC Completion (yyyy-mm-dd)',
        'Date of Repeat Viral Load - Post EAC VL Sample collected (yyyy-mm-dd)',
        'Repeat Viral load result (c/ml)- POST EAC',
        'Date of Repeat Viral load result- POST EAC VL', 'Date of devolvement',
        'Model devolved to', 'Date of current DSD', 'Current DSD model',
        'Current DSD Outlet', 'Date of Return of DSD Client to Facility (yyyy-mm-dd)',
        'Screening for Chronic Conditions', 'Co-morbidities',
        'Date of Cervical Cancer Screening (yyyy-mm-dd)',
        'Cervical Cancer Screening Type', 'Cervical Cancer Screening Method',
        'Result of Cervical Cancer Screening',
        'Date of Precancerous Lesions Treatment (yyyy-mm-dd)',
        'Precancerous Lesions Treatment Methods', 'Date Biometrics Enrolled (yyyy-mm-dd)',
        'Number of Fingers Captured', 'Date Biometrics Recapture (yyyy-mm-dd)',
        'Number of Fingers Recaptured', 'Case Manager'
    ]

    data = []
    
    if rows and len(rows) > 0:
        for index, row in enumerate(rows):
            row_data = []
            for col_idx, col in enumerate(columns):
                if col == 'S/No.':
                    row_data.append(index + 1)
                    continue
                
                mapping = {
                    'State': row.get('state'),
                    'L.G.A': row.get('lga'),
                    'LGA Of Residence': row.get('lgaofresidence'),
                    'Facility Name': row.get('facilityname'),
                    'DatimId': row.get('datimid'),
                    'Patient ID': row.get('personuuid') or row.get('uniquepersonuuid'),
                    'NDR Patient Identifier': row.get('ndrpatientidentifier'),
                    'Hospital Number': row.get('hospitalnumber'),
                    'Unique Id': row.get('uniqueid'),
                    'Household Unique No': row.get('householdnumber') or row.get('householduniqueno'),
                    'OVC Unique ID': row.get('ovcnumber') or row.get('ovcuniqueid'),
                    'Sex': row.get('gender'),
                    'Target group': row.get('targetgroup'),
                    'Current Weight (kg)': row.get('currentweight'),
                    'Pregnancy Status': row.get('pregnancystatus'),
                    'Date of Birth (yyyy-mm-dd)': row.get('dateofbirth'),
                    'Age': row.get('age'),
                    'Care Entry Point': row.get('careentry'),
                    'Date of Registration': row.get('dateofregistration'),
                    'Enrollment Date (yyyy-mm-dd)': row.get('dateofenrollment'),
                    'ART Start Date (yyyy-mm-dd)': row.get('artstartdate'),
                    'Last Pickup Date (yyyy-mm-dd)': row.get('lastpickupdate'),
                    'Months of ARV Refill': row.get('monthsofarxrefill') or row.get('monthsofarvrefill'),
                    'Regimen Line at ART Start': row.get('regimenlineatstart'),
                    'Regimen at ART Start': row.get('regimenatstart'),
                    'Date of Start of Current ART Regimen': row.get('dateofcurrentregimen'),
                    'Current Regimen Line': row.get('currentregimenline'),
                    'Current ART Regimen': row.get('currentartregimen'),
                    'Clinical Staging at Last Visit': row.get('currentclinicalstage'),
                    'Date of Last CD4 Count': row.get('dateoflastcd4count'),
                    'Last CD4 Count': row.get('lastcd4count'),
                    'Date of Viral Load Sample Collection (yyyy-mm-dd)': row.get('dateofviralloadsamplecollection'),
                    'Date of Current ViralLoad Result Sample (yyyy-mm-dd)': row.get('dateofcurrentviralloadsample'),
                    'Current Viral Load (c/ml)': row.get('currentviralload'),
                    'Date of Current Viral Load (yyyy-mm-dd)': row.get('dateofcurrentviralload'),
                    'Viral Load Indication': row.get('viralloadindication'),
                    'Viral Load Eligibility Status': row.get('vleligibilitystatus'),
                    'Date of Viral Load Eligibility Status': row.get('dateofvleligibilitystatus'),
                    'Current ART Status': row.get('currentstatus'),
                    'Date of Current ART Status': row.get('currentstatusdate'),
                    'Client Verification Outcome': row.get('clientverificationoutcome'),
                    'Cause of Death': row.get('causeofdeath'),
                    'VA Cause of Death': row.get('vacauseofdeath'),
                    'Previous ART Status': row.get('previousstatus'),
                    'Confirmed Date of Previous ART Status': row.get('previousstatusdate'),
                    'ART Enrollment Setting': row.get('enrollmentsetting'),
                    'Date of TB Screening (yyyy-mm-dd)': row.get('dateoftbscreened'),
                    'TB Screening Type': row.get('tbscreeningtype'),
                    'CAD Score': row.get('cadscore'),
                    'TB status': row.get('tbstatus'),
                    'Date of TB Sample Collection (yyyy-mm-dd)': row.get('dateoftbsamplecollection'),
                    'TB Diagnostic Test Type': row.get('tbdiagnostictesttype'),
                    'Date of TB Diagnostic Result Received (yyyy-mm-dd)': row.get('dateoftbdiagnosticresultreceived'),
                    'TB Diagnostic Result': row.get('tbdiagnosticresult'),
                    'Date of Additional TB Diagnosis Result using XRAY': row.get('datetbscorecad'),
                    'Additional TB Diagnosis Result using XRAY': row.get('resulttbscorecad'),
                    'Date of Start of TB Treatment (yyyy-mm-dd)': row.get('tbtreatmentstartdate'),
                    'TB Type (new, relapsed etc)': row.get('tbtreatementtype'),
                    'Date of Completion of TB Treatment (yyyy-mm-dd)': row.get('tbcompletiondate'),
                    'TB Treatment Outcome': row.get('tbtreatmentoutcome'),
                    'Date of TPT Start (yyyy-mm-dd)': row.get('dateofiptstart'),
                    'TPT Type': row.get('ipttype'),
                    'TPT Completion date (yyyy-mm-dd)': row.get('iptcompletiondate'),
                    'TPT Completion status': row.get('iptcompletionstatus'),
                    'Date of commencement of EAC (yyyy-mm-dd)': row.get('dateofcommencementofeac'),
                    'Number of EAC Sessions Completed': row.get('numberofeacsessioncompleted'),
                    'Date of last EAC Session Completed': row.get('dateoflasteacsessioncompleted'),
                    'Date of Extended EAC Completion (yyyy-mm-dd)': row.get('dateofextendeaccompletion'),
                    'Date of Repeat Viral Load - Post EAC VL Sample collected (yyyy-mm-dd)': row.get('dateofrepeatviralloadeacsamplecollection'),
                    'Repeat Viral load result (c/ml)- POST EAC': row.get('repeatviralloadresult'),
                    'Date of Repeat Viral load result- POST EAC VL': row.get('dateofrepeatviralloadresult'),
                    'Date of devolvement': row.get('dateofdevolvement'),
                    'Model devolved to': row.get('modeldevolvedto'),
                    'Date of current DSD': row.get('dateofcurrentdsd'),
                    'Current DSD model': row.get('currentdsdmodel'),
                    'Current DSD Outlet': row.get('currentdsdoutlet'),
                    'Date of Return of DSD Client to Facility (yyyy-mm-dd)': row.get('datereturntosite'),
                    'Screening for Chronic Conditions': None,
                    'Co-morbidities': None,
                    'Date of Cervical Cancer Screening (yyyy-mm-dd)': row.get('dateofcervicalcancerscreening'),
                    'Cervical Cancer Screening Type': row.get('cervicalcancerscreeningtype'),
                    'Cervical Cancer Screening Method': row.get('cervicalcancerscreeningmethod'),
                    'Result of Cervical Cancer Screening': row.get('resultofcervicalcancerscreening'),
                    'Date of Precancerous Lesions Treatment (yyyy-mm-dd)': row.get('treatmentmethoddate'),
                    'Precancerous Lesions Treatment Methods': row.get('cervicalcancertreatmentscreened'),
                    'Date Biometrics Enrolled (yyyy-mm-dd)': row.get('datebiometricsenrolled'),
                    'Number of Fingers Captured': row.get('numberoffingerscaptured'),
                    'Date Biometrics Recapture (yyyy-mm-dd)': row.get('datebiometricsrecaptured'),
                    'Number of Fingers Recaptured': row.get('numberoffingersrecaptured'),
                    'Case Manager': row.get('casemanager')
                }
                
                value = mapping.get(col)
                row_data.append(value if value is not None else '')
            
            data.append(row_data)
    else:
        data.append([''] * len(columns))

    df = pd.DataFrame(data, columns=columns)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='RADET Report', index=False)
        
        workbook = writer.book
        worksheet = writer.sheets['RADET Report']
        
        for i, col in enumerate(columns):
            column_letter = chr(65 + i) if i < 26 else chr(64 + i//26) + chr(65 + i%26)
            worksheet.column_dimensions[column_letter].width = min(40, max(len(col), 10))
        
        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    output.seek(0)
    return output.getvalue()

def generate_pharmacy_excel(rows):
    """Generate Pharmacy report Excel with exact column structure"""
    columns = [
        'S/No',
        'Facility Name',
        'DATIM Id',
        'Patient Id',
        'Hospital Num',
        'Date Visit(yyyy-mm-dd)',
        'Regimen Line',
        'Regimens(Include supported Drugs)',
        'Refill Period',
        'MMD_Type',
        'Next Appointment (yyyy-mm-dd)',
        'DSD Model'
    ]
    
    data = []
    
    if rows and len(rows) > 0:
        for index, row in enumerate(rows):
            # Format dates to remove time component if present
            date_visit = row.get('datevisit')
            if date_visit and ' ' in str(date_visit):
                date_visit = str(date_visit).split(' ')[0]
            elif date_visit and hasattr(date_visit, 'strftime'):
                date_visit = date_visit.strftime('%Y-%m-%d')
            
            next_appointment = row.get('nextappointment')
            if next_appointment and ' ' in str(next_appointment):
                next_appointment = str(next_appointment).split(' ')[0]
            elif next_appointment and hasattr(next_appointment, 'strftime'):
                next_appointment = next_appointment.strftime('%Y-%m-%d')
            
            row_data = [
                index + 1,  # S/No
                row.get('facilityname', ''),
                row.get('datimid', ''),
                row.get('patientid', ''),
                row.get('hospitalnum', ''),
                date_visit,
                row.get('regimenline', ''),
                row.get('regimens', ''),
                row.get('refillperiod', ''),
                row.get('mmdtype', ''),
                next_appointment,
                row.get('dsdmodel', '')
            ]
            data.append(row_data)
    else:
        # No data case - add empty row with headers only
        data.append([''] * len(columns))
    
    # Create DataFrame
    df = pd.DataFrame(data, columns=columns)
    
    # Create Excel file in memory
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='pharmacy-report', index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['pharmacy-report']
        
        # Set column widths based on content
        column_widths = {
            'A': 8,   # S/No
            'B': 35,  # Facility Name
            'C': 15,  # DATIM Id
            'D': 40,  # Patient Id
            'E': 18,  # Hospital Num
            'F': 22,  # Date Visit(yyyy-mm-dd)
            'G': 20,  # Regimen Line
            'H': 40,  # Regimens(Include supported Drugs)
            'I': 15,  # Refill Period
            'J': 12,  # MMD_Type
            'K': 25,  # Next Appointment (yyyy-mm-dd)
            'L': 15   # DSD Model
        }
        
        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width
        
        # Style headers
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply header styling
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Apply borders and alignment to data cells
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.border = thin_border
                if cell.column == 1:  # S/No column
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif cell.column in [6, 11]:  # Date columns
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Freeze the header row
        worksheet.freeze_panes = 'A2'
    
    output.seek(0)
    return output.getvalue()

# ============================================================
# API ENDPOINTS
# ============================================================

@app.route('/api/health', methods=['GET'])
def health_check():
    is_connected = connection_pool is not None
    return jsonify({
        'success': True,
        'status': 'online',
        'database': {
            'status': 'connected' if is_connected else 'disconnected',
            'configured': bool(db_config and db_config.get('database', {}).get('database')),
            'name': db_config.get('database', {}).get('database') if db_config else None
        },
        'timestamp': datetime.now().isoformat()
    })

@app.route('/api/config/test', methods=['POST'])
def test_connection():
    data = request.json
    try:
        conn = psycopg2.connect(
            host=data.get('host'),
            port=data.get('port'),
            database=data.get('database'),
            user=data.get('username'),
            password=data.get('password'),
            connect_timeout=5
        )
        cursor = conn.cursor()
        cursor.execute('SELECT NOW()')
        cursor.close()
        conn.close()
        
        return jsonify({'success': True, 'message': '✅ Connection successful!'})
    except Exception as error:
        return jsonify({'success': False, 'message': f'❌ Connection failed: {str(error)}'}), 500

@app.route('/api/config/save', methods=['POST'])
def save_config():
    global db_config
    config = request.json
    
    try:
        config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'config.json')
        with open(config_path, 'w') as f:
            json.dump(config, f, indent=2)
        
        db_config = config
        connect_to_database()
        
        return jsonify({'success': True, 'message': '✅ Configuration saved!'})
    except Exception as error:
        return jsonify({'success': False, 'message': str(error)}), 500

@app.route('/api/config/status', methods=['GET'])
def config_status():
    return jsonify({
        'configured': bool(db_config and db_config.get('database', {}).get('database')),
        'database': db_config.get('database', {}).get('database') if db_config else None,
        'host': db_config.get('database', {}).get('host') if db_config else None
    })

@app.route('/api/report/radet/generate', methods=['POST'])
def generate_radet_report():
    data = request.json
    start_date = data.get('startDate')
    end_date = data.get('endDate')
    row_limit = data.get('rowLimit', 100000)
    
    if not start_date or not end_date:
        return jsonify({'success': False, 'message': 'Start date and end date are required'}), 400
    
    if not connection_pool:
        return jsonify({'success': False, 'message': 'Database not configured'}), 400
    
    conn = None
    try:
        query = load_radet_query()
        if not query:
            return jsonify({'success': False, 'message': 'RADET query file not found'}), 404
        
        # Replace dates
        query = query.replace("'1980-01-01'", f"'{start_date}'")
        query = query.replace("CURRENT_DATE", f"'{end_date}'")
        
        # Add row limit if specified
        if row_limit > 0:
            query = f"{query} LIMIT {row_limit}"
        
        print(f'Executing RADET query for range: {start_date} to {end_date}')
        
        conn = connection_pool.getconn()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute(query)
        rows = cursor.fetchall()
        cursor.close()
        
        print(f'Query returned {len(rows)} rows')
        
        buffer = generate_radet_excel(rows)
        filename = f"RADET_Report_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}.xlsx"
        
        response = make_response(send_file(
            io.BytesIO(buffer),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        ))
        
        response.headers['Content-Length'] = len(buffer)
        return response
        
    except Exception as error:
        print(f'Query execution error: {str(error)}')
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error: {str(error)}'}), 500
    finally:
        if conn:
            connection_pool.putconn(conn)

@app.route('/api/report/pharmacy/generate', methods=['POST'])
def generate_pharmacy_report():
    data = request.json
    row_limit = data.get('rowLimit', 10000)
    
    if not connection_pool:
        return jsonify({'success': False, 'message': 'Database not configured'}), 400
    
    conn = None
    try:
        query = load_pharmacy_query()
        if not query:
            return jsonify({'success': False, 'message': 'Pharmacy query file not found'}), 404
        
        # Add row limit
        if row_limit > 0:
            query = f"{query} LIMIT {row_limit}"
        
        print(f'Executing Pharmacy query')
        
        conn = connection_pool.getconn()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute(query)
        rows = cursor.fetchall()
        cursor.close()
        
        print(f'Query returned {len(rows)} rows')
        
        buffer = generate_pharmacy_excel(rows)
        filename = f"Pharmacy_Report_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}.xlsx"
        
        response = make_response(send_file(
            io.BytesIO(buffer),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        ))
        
        response.headers['Content-Length'] = len(buffer)
        return response
        
    except Exception as error:
        print(f'Query execution error: {str(error)}')
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error: {str(error)}'}), 500
    finally:
        if conn:
            connection_pool.putconn(conn)

@app.route('/', defaults={'path': ''})
@app.route('/<path:path>')
def catch_all(path):
    if frontend_path and path and os.path.exists(os.path.join(frontend_path, path)):
        return app.send_static_file(path)
    elif frontend_path and os.path.exists(os.path.join(frontend_path, 'index.html')):
        return app.send_static_file('index.html')
    else:
        return jsonify({'error': 'Frontend not found'}), 404

# ============================================================
# START SERVER
# ============================================================
def start_server():
    load_config()
    connect_to_database()
    
    db_name = db_config.get('database', {}).get('database') if db_config else 'Not configured'
    db_status = '🟢 CONNECTED' if connection_pool else '🔴 DISCONNECTED'
    
    print(f"""
╔══════════════════════════════════════════════════════════════╗
║                                                              ║
║   🚀 SMART REPORT GENERATOR v4.0                            ║
║   ════════════════════════════════════════════════════════   ║
║                                                              ║
║   📡 Server:      http://localhost:{PORT}                      ║
║   🔧 Database:    {db_name}                                    ║
║   📊 Status:      {db_status}                                  ║
║   📋 Reports:     RADET (81 cols) | Pharmacy (13 cols)       ║
║                                                              ║
╚══════════════════════════════════════════════════════════════╝
    """)

if __name__ == '__main__':
    start_server()
    app.run(host='0.0.0.0', port=PORT, debug=False)